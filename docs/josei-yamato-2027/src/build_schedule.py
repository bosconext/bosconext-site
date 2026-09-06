# -*- coding: utf-8 -*-
"""
別紙1「スケジュール」生成スクリプト
------------------------------------------------------------------
★ 予定を直したいときは、下の ROWS の「●／◆」を書き換えてください。
   ●＝実施（継続）／◆＝重点（大会・イベント等の山場）
★ 実行:  python3 build_schedule.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

FONT = "游ゴシック"
NAVY, RED, GRAY, LINE = "1F3864", "B32424", "595959", "BFBFBF"
LIGHT, HEADFILL, BANDFILL = "DEEBF7", "EEF1F7", "F7F9FC"

# 2027年4月 ～ 2028年2月（11か月）
MONTHS = ["4月","5月","6月","7月","8月","9月","10月","11月","12月","1月","2月"]
M = {m: i for i, m in enumerate(MONTHS)}

def bar(marks):
    """marks: {"6月": "◆", ...} → 11要素のリスト"""
    row = [""] * len(MONTHS)
    for k, v in marks.items():
        row[M[k]] = v
    return row

ALL = {m: "●" for m in MONTHS}

# 区分, 実施項目, 回数・数量, ガント, 備考
ROWS = [
    ("準　備", "事業計画の確定・連携先との調整", "―",
     bar({"4月":"●","5月":"●"}), "連携機関への依頼・内諾の取得"),
    ("", "会場の確保（抽選申込・予約）", "年52回分",
     bar(ALL), "区立施設の抽選は原則2か月前。継続的に申込む"),
    ("", "傷害保険の加入手続", "会員20名",
     bar({"4月":"●"}), "スポーツ安全保険"),
    ("", "用具・救急用品の調達", "一式",
     bar({"4月":"●","5月":"●"}), "ボール・ビブス・トレーニング用具ほか"),

    ("柱1\nオープン\n体験会", "広報物（チラシ）の制作・配布", "2,000部",
     bar({"5月":"●","8月":"●","11月":"●"}), "連携機関・地域の窓口へ配架"),
    ("", "第1回 オープン体験会・交流会", "20〜30名",
     bar({"6月":"◆"}), ""),
    ("", "第2回 オープン体験会・交流会", "20〜30名",
     bar({"8月":"◆"}), ""),
    ("", "第3回 オープン体験会・交流会", "20〜30名",
     bar({"10月":"◆"}), ""),
    ("", "第4回 オープン体験会・交流会", "20〜30名",
     bar({"12月":"◆"}), ""),
    ("", "第5回 オープン体験会・交流会", "20〜30名",
     bar({"2月":"◆"}), "活動報告会と同日開催も可"),
    ("", "参加者アンケートの集計・振り返り", "各回終了後",
     bar({"7月":"●","9月":"●","11月":"●","1月":"●","2月":"●"}), "次回プログラムへ反映"),

    ("柱2\n練習・\n大会", "定例練習（毎週土曜日）", "年40回",
     bar(ALL), "大田区・目黒区・品川区ほか"),
    ("", "強化練習会（外部指導者の招へい）", "年12回",
     bar(ALL), "月1回。傷害予防と戦術面の向上"),
    ("", "ソーシャルフットボール東京都大会", "出場",
     bar({"5月":"◆","6月":"●"}), "開催時期は主催者の発表により変動"),
    ("", "全国大会 予選・関東大会", "出場",
     bar({"7月":"◆"}), "例年7月開催（2026年は7月19日）"),
    ("", "JIFF ソーシャルフットボール全国大会", "出場（予選通過時）",
     bar({"11月":"◆"}), "例年11月開催（2026年は11月21日）"),
    ("", "交流試合・練習試合", "随時",
     bar({"6月":"●","9月":"●","12月":"●","2月":"●"}), "近隣チーム・連携機関との交流"),

    ("柱3\n普及・\n啓発", "ハンドブックの企画・原稿執筆", "A5判24頁",
     bar({"6月":"●","7月":"●","8月":"●"}), "当事者・家族・支援者・運営者向けの4部構成"),
    ("", "編集・デザイン・校正", "―",
     bar({"9月":"●","10月":"●"}), "当事者メンバーによる読み合わせを実施"),
    ("", "印刷", "400部",
     bar({"11月":"◆"}), ""),
    ("", "配布（医療機関・福祉事業所・自治体ほか）", "50か所以上",
     bar({"12月":"●","1月":"●","2月":"●"}), "手渡しでの説明を基本とする"),
    ("", "ウェブサイト・SNS・noteでの発信", "随時",
     bar(ALL), ""),
    ("", "活動報告会（オープン形式）", "1回",
     bar({"2月":"◆"}), "参加者・連携機関・地域の方を招く"),

    ("評価・\n報告", "事業評価（アンケート分析・ヒアリング）", "―",
     bar({"1月":"●","2月":"●"}), "アウトカム指標の達成状況を確認"),
    ("", "会計整理・領収書の整理", "―",
     bar({"12月":"●","1月":"●","2月":"●"}), ""),
    ("", "実績報告書の作成・財団へ提出", "―",
     bar({"2月":"◆"}), "2028年2月29日までに事業完了"),
]

wb = Workbook()
ws = wb.active
ws.title = "スケジュール"

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

NCOL = 3 + len(MONTHS) + 1          # 区分/項目/数量 + 11か月 + 備考 = 15
LAST = get_column_letter(NCOL)

# ---- 標題 ----
ws["A1"] = "別紙1　スケジュール"
ws["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
ws.merge_cells("A1:%s1" % LAST)

ws["A2"] = "事業名：「どんな壁も、乗り越えられる。」ソーシャルフットボールによる精神障がい当事者の健康づくり・社会参加促進事業"
ws["A2"].font = Font(name=FONT, size=10)
ws.merge_cells("A2:%s2" % LAST)

ws["A3"] = "申請団体：BOSCO NEXT（ボスコ ネクスト）　／　実施期間：2027年4月1日 〜 2028年2月29日（11か月）"
ws["A3"].font = Font(name=FONT, size=10)
ws.merge_cells("A3:%s3" % LAST)

HDR1, HDR2, FIRST = 5, 6, 7

# ---- 見出し（2段） ----
ws.cell(HDR1, 1, "区　分"); ws.cell(HDR1, 2, "実 施 項 目"); ws.cell(HDR1, 3, "回数・数量")
ws.cell(HDR1, 3 + len(MONTHS) + 1, "備　考")
for c in (1, 2, 3, NCOL):
    ws.merge_cells(start_row=HDR1, start_column=c, end_row=HDR2, end_column=c)

ws.cell(HDR1, 4, "2027年")
ws.merge_cells(start_row=HDR1, start_column=4, end_row=HDR1, end_column=4 + 8)   # 4月〜12月
ws.cell(HDR1, 13, "2028年")
ws.merge_cells(start_row=HDR1, start_column=13, end_row=HDR1, end_column=14)     # 1月〜2月

for i, m in enumerate(MONTHS):
    ws.cell(HDR2, 4 + i, m)

for r in (HDR1, HDR2):
    for c in range(1, NCOL + 1):
        cell = ws.cell(r, c)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box

# ---- 本体 ----
r = FIRST
for kubun, item, qty, gantt, memo in ROWS:
    ws.cell(r, 1, kubun).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(r, 1).font = Font(name=FONT, size=9, bold=True, color=NAVY)
    ws.cell(r, 2, item).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(r, 3, qty).alignment = Alignment(horizontal="center", vertical="center")
    for i, v in enumerate(gantt):
        ws.cell(r, 4 + i, v).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(r, NCOL, memo).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for c in range(1, NCOL + 1):
        cell = ws.cell(r, c)
        if cell.font.name != FONT or c != 1:
            cell.font = Font(name=FONT, size=9)
        cell.border = box
    ws.cell(r, 1).font = Font(name=FONT, size=9, bold=True, color=NAVY)
    if kubun:                       # 区分の切り替わり行に上罫線を強調
        for c in range(1, NCOL + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, bottom=thin,
                                          top=Side(style="medium", color=NAVY))
    ws.row_dimensions[r].height = 21
    r += 1
LASTROW = r - 1

# 区分セルを縦に結合
start = FIRST
for i in range(len(ROWS) + 1):
    row = FIRST + i
    is_new = (i == len(ROWS)) or bool(ROWS[i][0])
    if is_new and row > start:
        if row - 1 > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        start = row

# ---- 条件付き書式（●／◆ を入力すると自動で色がつく） ----
grange = "D%d:N%d" % (FIRST, LASTROW)
ws.conditional_formatting.add(grange, FormulaRule(
    formula=['D%d="◆"' % FIRST], stopIfTrue=True,
    fill=PatternFill("solid", start_color=NAVY, end_color=NAVY),
    font=Font(name=FONT, size=9, bold=True, color="FFFFFF")))
ws.conditional_formatting.add(grange, FormulaRule(
    formula=['LEN(D%d)>0' % FIRST],
    fill=PatternFill("solid", start_color=LIGHT, end_color=LIGHT),
    font=Font(name=FONT, size=9, color="2E75B6")))

# ---- 凡例・注記 ----
r = LASTROW + 2
notes = [
    ("【凡例】　●＝実施（継続して取り組む月）　　◆＝重点（大会・イベント等、その月の山場）", True),
    ("※ セルに「●」または「◆」を入力すると、色は自動でつきます。予定を変更する場合は文字を書き換えてください。", False),
    ("※ 大会の開催時期は主催者の発表により変動します。確定次第、本表を更新します。", False),
    ("※ 事業は2027年4月1日に開始し、2028年2月29日までにすべて完了します。", False),
]
for text, bold in notes:
    ws.cell(r, 1, text).font = Font(name=FONT, size=9, bold=bold, color=("222222" if bold else GRAY))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

# ---- 列幅・印刷設定 ----
ws.column_dimensions["A"].width = 8.5
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 14
for i in range(len(MONTHS)):
    ws.column_dimensions[get_column_letter(4 + i)].width = 5.4
ws.column_dimensions[LAST].width = 32
ws.row_dimensions[1].height = 24
ws.row_dimensions[HDR1].height = 18
ws.row_dimensions[HDR2].height = 18
ws.freeze_panes = "D%d" % FIRST

ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = "%d:%d" % (HDR1, HDR2)
ws.print_area = "A1:%s%d" % (LAST, r - 1)
ws.page_margins.left = ws.page_margins.right = 0.4
ws.page_margins.top = ws.page_margins.bottom = 0.5

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "02_スケジュール.xlsx")
wb.save(out)
print("wrote", os.path.normpath(out))
