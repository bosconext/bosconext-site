# -*- coding: utf-8 -*-
"""
別紙2「費用積算表」生成スクリプト
------------------------------------------------------------------
★ 金額を直したいときは、下の GROUPS / INCOME を書き換えてください。
   （単価・数量を変えれば、金額・小計・合計はExcelの数式で自動計算されます）
★ 実行:  python3 build_budget.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "游ゴシック"
NAVY, GRAY, LINE = "1F3864", "595959", "BFBFBF"
SUBFILL, HEADFILL, YELLOW = "EEF1F7", "F7F9FC", "FFF2CC"
NUM = '#,##0;-#,##0;"－"'

# 費目区分 → [(費目, 内訳・積算根拠, 単価, 数量, 単位, うち助成金申請額)]
GROUPS = [
 ("1", "会場使用料", [
   ("オープン体験会 会場使用料", "体育館・フットサルコート（3時間／回）　22,000円 × 5回", 22000, 5, "回", 110000),
   ("強化練習会 会場使用料",     "体育館（3時間／回）　15,000円 × 12回（月1回）",           15000, 12, "回", 180000),
   ("定例練習 会場使用料",       "体育館・コート（2時間／回）　8,000円 × 40回（毎週土曜）",   8000, 40, "回",      0),
 ]),
 ("2", "謝　金", [
   ("外部指導者 謝金",           "フットサル指導者　15,000円 × 10回（強化練習会・体験会）",  15000, 10, "回", 150000),
   ("専門職アドバイザー 謝金",   "精神保健福祉士等　10,000円 × 5回（相談対応・安全管理）",   10000,  5, "回",  50000),
 ]),
 ("3", "保険料", [
   ("スポーツ安全保険",          "会員分　1,850円 × 20名（年額）",                            1850, 20, "名",  37000),
   ("行事参加者傷害保険",        "オープン体験会の参加者分　300円 × 100名",                    300,100, "名",  30000),
 ]),
 ("4", "用具・消耗品費", [
   ("フットサルボール",          "検定球　6,000円 × 10個",                                    6000, 10, "個",  60000),
   ("ビブス（体験会用）",        "1,200円 × 30枚",                                            1200, 30, "枚",  36000),
   ("トレーニング用具一式",      "マーカーコーン・ラダー・ミニゴール等　30,000円 × 1式",      30000,  1, "式",  30000),
   ("救急・衛生用品一式",        "アイシング用品、消耗品、体調確認シート等　20,000円 × 1式",  20000,  1, "式",  20000),
   ("ゴレイロ用具",              "グローブ・プロテクター等　25,000円 × 1式",                  25000,  1, "式",  25000),
 ]),
 ("5", "旅費・交通費", [
   ("全国大会 遠征交通費",       "選手・スタッフ　3,000円 × 16名",                             3000, 16, "名",  48000),
   ("体験会 運営スタッフ交通費", "1,000円 × 4名 × 5回",                                       1000, 20, "名回", 20000),
 ]),
 ("6", "広報・普及啓発費", [
   ("チラシ・ポスター印刷費",    "A4カラー　15円 × 2,000部",                                     15,2000,"部",  30000),
   ("ハンドブック制作・印刷費",  "A5判24頁　250円 × 400部",                                     250, 400, "部", 100000),
   ("記録・広報用機材",          "デジタルカメラ　40,000円 × 1台（活動記録・報告書用）",       40000,  1, "台",  40000),
 ]),
 ("7", "事務費", [
   ("通信運搬費",                "ハンドブック発送・郵送料　12,000円 × 1式",                  12000,  1, "式",  12000),
   ("事務用品・書類印刷費",      "10,000円 × 1式",                                            10000,  1, "式",  10000),
   ("報告書作成費",              "データ集計・印刷　10,000円 × 1式",                          10000,  1, "式",  10000),
   ("振込手数料等",              "2,000円 × 1式",                                              2000,  1, "式",   2000),
 ]),
 ("8", "その他（自己資金で負担）", [
   ("大会参加費",                "東京都大会・関東大会・全国大会　10,000円 × 3大会",          10000,  3, "大会",     0),
   ("ユニフォーム・チーム備品の更新", "10,000円 × 1式",                                       10000,  1, "式",       0),
 ]),
]

# 収入の部 …（費目, 内訳, 金額, 要記入か）
INCOME = [
  ("ヤマト福祉財団 障がい者福祉助成金", "本申請額", 1000000, False),
  ("会　費",                           "1,500円 × 16名 × 12か月",  288000, True),
  ("オープン体験会 参加費",            "500円 × 100名",              50000, True),
  ("寄付金・その他",                   "―",                          22000, True),
]

wb = Workbook()
ws = wb.active
ws.title = "費用積算表"

thin = Side(style="thin", color=LINE)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
NCOL, LAST = 9, "I"

def style(r, c, *, size=9, bold=False, color="222222", fill=None,
          align="left", wrap=False, fmt=None, border=True):
    cell = ws.cell(r, c)
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = box
    return cell

# ---- 標題 ----
ws["A1"] = "別紙2　費用積算表"
ws["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
ws.merge_cells("A1:%s1" % LAST)
ws["A2"] = "事業名：「どんな壁も、乗り越えられる。」ソーシャルフットボールによる精神障がい当事者の健康づくり・社会参加促進事業"
ws["A2"].font = Font(name=FONT, size=10)
ws.merge_cells("A2:%s2" % LAST)
ws["A3"] = "申請団体：BOSCO NEXT（ボスコ ネクスト）　／　実施期間：2027年4月1日 〜 2028年2月29日"
ws["A3"].font = Font(name=FONT, size=10)
ws.merge_cells("A3:%s3" % LAST)

r = 5
ws.cell(r, 1, "【支出の部】")
ws.cell(r, 1).font = Font(name=FONT, size=11, bold=True, color=NAVY)
ws.cell(r, NCOL, "（単位：円）")
ws.cell(r, NCOL).font = Font(name=FONT, size=9, color=GRAY)
ws.cell(r, NCOL).alignment = Alignment(horizontal="right")
r += 1

HDR = r
for c, txt in enumerate(
    ["No.", "費　目", "内訳・積算根拠", "単　価", "数量", "単位", "金　額",
     "うち\n助成金申請額", "自己資金"], start=1):
    style(HDR, c, size=9, bold=True, color="FFFFFF", fill=NAVY, align="center", wrap=True)
    ws.cell(HDR, c, txt)
ws.row_dimensions[HDR].height = 30
r += 1

sub_rows = []
for gno, gname, items in GROUPS:
    first = r
    for i, (name, basis, price, qty, unit, grant) in enumerate(items, start=1):
        ws.cell(r, 1, "%s-%d" % (gno, i));                style(r, 1, align="center")
        ws.cell(r, 2, name);                              style(r, 2, wrap=True)
        ws.cell(r, 3, basis);                             style(r, 3, size=8.5, wrap=True)
        ws.cell(r, 4, price);                             style(r, 4, align="right", fmt=NUM)
        ws.cell(r, 5, qty);                               style(r, 5, align="center", fmt='#,##0')
        ws.cell(r, 6, unit);                              style(r, 6, align="center")
        ws.cell(r, 7, "=D%d*E%d" % (r, r));               style(r, 7, align="right", fmt=NUM)
        ws.cell(r, 8, grant);                             style(r, 8, align="right", fmt=NUM)
        ws.cell(r, 9, "=G%d-H%d" % (r, r));               style(r, 9, align="right", fmt=NUM)
        ws.row_dimensions[r].height = 22
        r += 1
    last = r - 1
    ws.cell(r, 1, gno);                                   style(r, 1, bold=True, align="center", fill=SUBFILL, color=NAVY)
    ws.cell(r, 2, "%s　小計" % gname);                    style(r, 2, bold=True, fill=SUBFILL, color=NAVY)
    for c in (3, 4, 5, 6):
        style(r, c, fill=SUBFILL)
    for c, col in ((7, "G"), (8, "H"), (9, "I")):
        ws.cell(r, c, "=SUM(%s%d:%s%d)" % (col, first, col, last))
        style(r, c, bold=True, align="right", fmt=NUM, fill=SUBFILL, color=NAVY)
    sub_rows.append(r)
    ws.row_dimensions[r].height = 20
    r += 1

TOTAL = r
ws.cell(TOTAL, 1, "合計"); ws.merge_cells(start_row=TOTAL, start_column=1, end_row=TOTAL, end_column=6)
for c in range(1, 7):
    style(TOTAL, c, size=11, bold=True, color="FFFFFF", fill=NAVY, align="center")
ws.cell(TOTAL, 1, "総 事 業 費　／　助 成 申 請 額　／　自 己 資 金")
for c, col in ((7, "G"), (8, "H"), (9, "I")):
    ws.cell(TOTAL, c, "=" + "+".join("%s%d" % (col, x) for x in sub_rows))
    style(TOTAL, c, size=11, bold=True, color="FFFFFF", fill=NAVY, align="right", fmt=NUM)
ws.row_dimensions[TOTAL].height = 26
r = TOTAL + 2

# ---- 収入の部 ----
ws.cell(r, 1, "【収入の部】")
ws.cell(r, 1).font = Font(name=FONT, size=11, bold=True, color=NAVY)
ws.cell(r, NCOL, "（単位：円）")
ws.cell(r, NCOL).font = Font(name=FONT, size=9, color=GRAY)
ws.cell(r, NCOL).alignment = Alignment(horizontal="right")
r += 1

IHDR = r
ws.cell(IHDR, 1, "No."); ws.cell(IHDR, 2, "収入の内容")
ws.cell(IHDR, 3, "内訳・算出根拠"); ws.cell(IHDR, 7, "金　額")
ws.merge_cells(start_row=IHDR, start_column=3, end_row=IHDR, end_column=6)
ws.merge_cells(start_row=IHDR, start_column=7, end_row=IHDR, end_column=9)
for c in range(1, NCOL + 1):
    style(IHDR, c, size=9, bold=True, color="FFFFFF", fill=NAVY, align="center")
r += 1

ifirst = r
for i, (name, basis, amount, editable) in enumerate(INCOME, start=1):
    ws.cell(r, 1, i);      style(r, 1, align="center")
    ws.cell(r, 2, name);   style(r, 2, wrap=True)
    ws.cell(r, 3, basis);  style(r, 3, size=8.5, wrap=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for c in (4, 5, 6):
        style(r, c)
    ws.cell(r, 7, amount)
    style(r, 7, align="right", fmt=NUM, fill=(YELLOW if editable else None))
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    for c in (8, 9):
        style(r, c, fill=(YELLOW if editable else None))
    ws.row_dimensions[r].height = 20
    r += 1
ilast = r - 1

ws.cell(r, 1, "収入合計"); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
for c in range(1, 7):
    style(r, c, size=10, bold=True, color="FFFFFF", fill=NAVY, align="center")
ws.cell(r, 1, "収 入 合 計")
ws.cell(r, 7, "=SUM(G%d:G%d)" % (ifirst, ilast))
style(r, 7, size=10, bold=True, color="FFFFFF", fill=NAVY, align="right", fmt=NUM)
ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
for c in (8, 9):
    style(r, c, fill=NAVY)
INCTOTAL = r
r += 1

ws.cell(r, 1, "差引（収入合計 － 総事業費）　※0になることを確認してください")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
style(r, 1, size=9, bold=True, color=NAVY, fill=SUBFILL, align="left")
for c in range(2, 7):
    style(r, c, fill=SUBFILL)
ws.cell(r, 7, "=G%d-G%d" % (INCTOTAL, TOTAL))
style(r, 7, size=9, bold=True, align="right", fmt='#,##0;-#,##0;0', fill=SUBFILL, color=NAVY)
ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
for c in (8, 9):
    style(r, c, fill=SUBFILL)
r += 2

# ---- 注記 ----
for text, bold in [
    ("【積算の前提と留意事項】", True),
    ("① 単価は2026年9月時点の実勢価格および当団体の過去の支出実績に基づく見込みです。契約時に変動する場合があります。", False),
    ("② 金額はすべて消費税込みです。", False),
    ("③ 黄色のセル（収入の部）は、貴団体の実情に合わせてご記入・ご確認ください。単価・数量を書き換えると、金額・小計・合計は自動で計算されます。", False),
    ("④ 定例練習の会場使用料（320,000円）、大会参加費、ユニフォーム・備品更新費は、会費等の自己資金で負担します。", False),
    ("⑤ 財団の募集要項において助成対象外とされる費目があった場合は、当該費目を自己資金へ振り替えて調整し、助成申請額は1,000,000円を上限とします。", False),
    ("⑥ 助成決定後は、費目ごとに領収書を保管し、実績報告書とあわせてご報告いたします。", False),
]:
    ws.cell(r, 1, text)
    ws.cell(r, 1).font = Font(name=FONT, size=9, bold=bold, color=("222222" if bold else GRAY))
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    r += 1
LASTROW = r - 1

# ---- 列幅・印刷設定 ----
for col, w in zip("ABCDEFGHI", [6, 22, 33, 9.5, 6, 6, 11.5, 12.5, 11]):
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 24
ws.freeze_panes = "A%d" % (HDR + 1)

ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_area = "A1:%s%d" % (LAST, LASTROW)
ws.page_margins.left = ws.page_margins.right = 0.4
ws.page_margins.top = ws.page_margins.bottom = 0.5
wb.calculation.fullCalcOnLoad = True

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "03_費用積算表.xlsx")
wb.save(out)
print("wrote", os.path.normpath(out))
